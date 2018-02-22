import htmlPy
import json
import sqlite3
import os
from shutil import copyfile
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
oscwd = os.getcwd()


#Copy tempalate workbook
copyfile("ScoutingData_template.xlsx","ScoutingData.xlsx")
wb = load_workbook('ScoutingData.xlsx')

#Connect To Database
db = sqlite3.connect('CombinedMatchedData')
c = db.cursor()
#Read In Team List

c.execute("SELECT DISTINCT teamName FROM matchdata")

rows = c.fetchall()

for row in rows:
    count = 0
    for col in row:
        print(row[count])
        ws = wb.create_sheet(row[count])
        count = count + 1


# Get Team Match Data

#Put data in excel



ws['A4'] = 573


wb.save('ScoutingData.xlsx')
db.close()

#Add in Averages Max Mins
#Add in Summary Sheet

    
    

        
