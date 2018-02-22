#Opens a template workbook and creates a copy of a template worksheet
#within a copy of the workbook

#Import libraries
import os
from openpyxl import load_workbook
import openpyxl
import shutil
from datetime import datetime
import json
import sqlite3

#Print Script Start
print '!!!!!!!!! Script Start !!!!!!!!!'

#Changes Working Directory To a Known Location
#os.chdir('C:\Users\es2433\Desktop\PV_PythonSandbox\ExcelBookLayout')

#Copy Excel file from template file with datetime stamp
CurrDateTime = str(datetime.now())
CurrDateTime = CurrDateTime.replace(':','-')
CurrDateTime = CurrDateTime.replace(' ','_')
CurrDateTime = CurrDateTime.replace('.','-')
shutil.copy('ScoutingData_template.xlsx','ScoutingData_'+CurrDateTime+'.xlsx')

#Connect To Match Database
db = sqlite3.connect('CombinedMatchedData')
c = db.cursor()

#Opens Excel Spreadsheet
wb = load_workbook('ScoutingData_'+CurrDateTime+'.xlsx')

#Find template worksheet by name
ws_template = wb['template']#.get_sheet_by_name(name = 'template')

# ----------------------------------------
#Random code isn't needed

##my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
##my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
##ws1["A1"].value = 'BLAH'
##ws1["A1"].fill = my_fill

#----------------------------------------

#Read in team numbers from team list
wsteam_list = wb['team_list']#.get_sheet_by_name(name = 'team_list')
TeamList=[]
TeamNameList=[]

#Populate Team List Based Upon Excel Data
for rowindex in range(wsteam_list.max_row):
    TeamList.append(str(wsteam_list.cell(row = rowindex + 1,column = 1).value))
    TeamNameList.append(str(wsteam_list.cell(row = rowindex + 1,column = 2).value))

#Loop through team list creating worksheet for each one
for inc, team in enumerate(TeamList):
#Create copy of template sheet
    ws2 = wb.copy_worksheet(ws_template)

#Rename Newly Copied Sheet
    ws2.title = team
#----------------------------------------

    #Fill out data on sheet
    ws2.cell(column=2,row=1).value = team
    ws2.cell(column=2,row=2).value = TeamNameList[inc]
    
    #Put in pit scouting data

    #Put in match scouting data
    c.execute("SELECT * FROM matchdata WHERE teamName="+team)
    rows = c.fetchall()
    for inc2, row in enumerate(rows):
        #crosslineauto 0
        #exchangeauto 1
        #matchName 2
        #notes 3
        #scoutName 4
        #deployRamp 5
        #startingposauto 6
        #driveOnPlatform 7
        #teamName 8
        #switchclosetele 9
        #allianceStation 10
        #switchauto 11
        #driveOnRamp 12
        #climbWithOther 13
        #scaleauto 14
        #scaletele 15
        #climb 16
        #exchangetele 17
        #switchfartele 18
        
        #somedata cleanup
        if row[11]=='':
            row11='0'
        else:
            row11 = row[11]
        if row[14]=='':
            row14='0'
        else:
            row14 = row[14]
        if row[1]=='':
            row1='0'
        else:
            row1 = row[1]
        if row[9]=='':
            row9='0'
        else:
            row9 = row[9]
        if row[15]=='':
            row15='0'
        else:
            row15 = row[15]
        if row[18]=='':
            row18='0'
        else:
            row18 = row[18]
        if row[17]=='':
            row17='0'
        else:
            row17 = row[17]
        #Put Data into proper cells on worksheet
        ws2.cell(row=inc2+17,column=1).value = row[2]
        ws2.cell(row=inc2+17,column=2).value = row[10]
        ws2.cell(row=inc2+17,column=3).value = row[4]
        
        ws2.cell(row=inc2+17,column=4).value = int(row11)
        ws2.cell(row=inc2+17,column=5).value = int(row14)
        ws2.cell(row=inc2+17,column=6).value = int(row1)
        ws2.cell(row=inc2+17,column=7).value = row[0]
        ws2.cell(row=inc2+17,column=8).value = row[6]
        ws2.cell(row=inc2+17,column=9).value = int(row9)
        ws2.cell(row=inc2+17,column=10).value = int(row15)
        ws2.cell(row=inc2+17,column=11).value = int(row18)
        ws2.cell(row=inc2+17,column=12).value = int(row17)
        ws2.cell(row=inc2+17,column=13).value = row[7]
        ws2.cell(row=inc2+17,column=14).value = row[12]
        ws2.cell(row=inc2+17,column=15).value = row[5]
        ws2.cell(row=inc2+17,column=16).value = row[16]
        ws2.cell(row=inc2+17,column=17).value = row[13]
        ws2.cell(row=inc2+17,column=18).value = row[3]

###Put in picture ----------------------------------
##    if (os.path.isfile('Pic/'+team+'.jpg')): #Check to make sure picture exists
##        teamnum = team
##    else:
##        teamnum ='WILogo' #Stand in Photo For Teams With Missing Picture
##
##    img = openpyxl.drawing.image.Image('Pic/'+teamnum+'.jpg')
##    img.anchor(ws2['G1'])
##    ws2.add_image(img)
# ------------------------------------------------------

    # Add in Summary Page Info
    ws_sum = wb['Summary']#.get_sheet_by_name(name = 'Summary')

    #Adds information to summary sheet from team worksheet
    ws_sum.cell(row = inc+9,column = 1).value = "='"+team+"'!"+"B1" #Team Number

    #AVG
    ws_sum.cell(row = inc+9,column = 2).value = "='"+team+"'!"+"D14" #Autos
    ws_sum.cell(row = inc+9,column = 3).value = "='"+team+"'!"+"E14"
    ws_sum.cell(row = inc+9,column = 4).value = "='"+team+"'!"+"F14"
    ws_sum.cell(row = inc+9,column = 5).value = "='"+team+"'!"+"G14"

    ws_sum.cell(row = inc+9,column = 6).value = "='"+team+"'!"+"I14" #Tele
    ws_sum.cell(row = inc+9,column = 7).value = "='"+team+"'!"+"J14"
    ws_sum.cell(row = inc+9,column = 8).value = "='"+team+"'!"+"K14"
    ws_sum.cell(row = inc+9,column = 9).value = "='"+team+"'!"+"L14"

    ws_sum.cell(row = inc+9,column = 10).value = "='"+team+"'!"+"M14" #Endgame  
    ws_sum.cell(row = inc+9,column = 11).value = "='"+team+"'!"+"N14"
    ws_sum.cell(row = inc+9,column = 12).value = "='"+team+"'!"+"O14"
    ws_sum.cell(row = inc+9,column = 13).value = "='"+team+"'!"+"P14"
    ws_sum.cell(row = inc+9,column = 14).value = "='"+team+"'!"+"Q14"

    #MAX
    ws_sum.cell(row = inc+9,column = 15).value = "='"+team+"'!"+"D15" #Autos
    ws_sum.cell(row = inc+9,column = 16).value = "='"+team+"'!"+"E15"
    ws_sum.cell(row = inc+9,column = 17).value = "='"+team+"'!"+"F15"
    
    ws_sum.cell(row = inc+9,column = 18).value = "='"+team+"'!"+"I15" #Tele
    ws_sum.cell(row = inc+9,column = 19).value = "='"+team+"'!"+"J15"
    ws_sum.cell(row = inc+9,column = 20).value = "='"+team+"'!"+"K15"
    ws_sum.cell(row = inc+9,column = 21).value = "='"+team+"'!"+"L15"

    #MIN
    ws_sum.cell(row = inc+9,column = 22).value = "='"+team+"'!"+"D16" #Autos
    ws_sum.cell(row = inc+9,column = 23).value = "='"+team+"'!"+"E16"
    ws_sum.cell(row = inc+9,column = 24).value = "='"+team+"'!"+"F16"

    ws_sum.cell(row = inc+9,column = 25).value = "='"+team+"'!"+"I16" #Tele
    ws_sum.cell(row = inc+9,column = 26).value = "='"+team+"'!"+"J16"
    ws_sum.cell(row = inc+9,column = 27).value = "='"+team+"'!"+"K16"
    ws_sum.cell(row = inc+9,column = 28).value = "='"+team+"'!"+"L16"
    #ws["A1"] = "=SUM('1'!B1 + '2'!B1 )" #Example
    
#Save workbook
wb.save('ScoutingData_'+CurrDateTime+'.xlsx')

#Print Script End
print '!!!!!!!!! File Created !!!!!!!!!'

